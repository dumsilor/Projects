#!/usr/bin/env python3

import shutil
import openpyxl
import datetime
import os


os.chdir("/shared/Bandwidth Report/Scripts")

print("Putting all the values in NovoCom bandwidth report.........")

uk_in  = None
uk_out  = None
bsccl_in = None
bsccl_out = None
airtel_in = None
airtel_out = None
sg_SMW4_in = None
sg_SMW4_out = None
fna_in = None
fna_out = None
akamai_in = None
akamai_out = None
singtel_in = None
singtel_out = None
TATA_in =  None
TATA_out =None
digicon_in = None
digicon_out = None
Novotel_IPLC_in  = None
Novotel_IPLC_out = None
Novotel_SG_SW_in = None
Novotel_SG_SW_out = None
UIL_SG_in = None
UIL_SG_out = None
level3_in = None
level3_out = None
nv_uk_in = None
nv_uk_out = None
dbl_in = None
dbl_out= None
google_node_1_in = None
google_node_1_out = None
google_node_2_in = None
google_node_2_out = None
IC_novocom_in =None
IC_novocom_out = None
BDIX_in = None
BDIX_out = None

raw_bandwidth_file =  openpyxl.load_workbook("nv_raw_bw.xlsx")
raw_sheet = None
raw_sheets = raw_bandwidth_file.sheetnames
for item in raw_sheets:
    raw_sheet = raw_bandwidth_file[item]
    bsccl_in = raw_sheet["B1"].value
    bsccl_out = raw_sheet["C1"].value
    airtel_in = raw_sheet["B2"].value
    airtel_out = raw_sheet["C2"].value
    uk_in = raw_sheet["B3"].value
    uk_out = raw_sheet["C3"].value
    sg_SMW4_in = raw_sheet["B4"].value
    sg_SMW4_out = raw_sheet["C4"].value
    fna_in = raw_sheet["B5"].value
    fna_out = raw_sheet["C5"].value
    akamai_in = raw_sheet["B6"].value
    akamai_out = raw_sheet["C6"].value
    singtel_in = raw_sheet["B7"].value
    singtel_out = raw_sheet["C7"].value
    TATA_in =  raw_sheet["B8"].value
    TATA_out =raw_sheet["C8"].value
    digicon_in = raw_sheet["B9"].value
    digicon_out = raw_sheet["C9"].value
    Novotel_IPLC_in  = raw_sheet["B10"].value
    Novotel_IPLC_out = raw_sheet["C10"].value
    Novotel_SG_SW_in = raw_sheet["B11"].value
    Novotel_SG_SW_out = raw_sheet["C11"].value
    UIL_SG_in = raw_sheet["B12"].value
    UIL_SG_out = raw_sheet["C12"].value
    level3_in = raw_sheet["B13"].value
    level3_out = raw_sheet["C13"].value
    nv_uk_in = raw_sheet["B14"].value
    nv_uk_out = raw_sheet["C14"].value
    dbl_in = raw_sheet["B15"].value
    dbl_out= raw_sheet["C15"].value
    google_node_1_in = raw_sheet["B16"].value
    google_node_1_out = raw_sheet["C16"].value
    google_node_2_in = raw_sheet["B17"].value
    google_node_2_out = raw_sheet["C17"].value
raw_bandwidth_file.close()

file =  openpyxl.load_workbook("../NovoCom-BW-Report.xlsx")
sheets = file.sheetnames
sheet = None

ITC_sheet = file[sheets[0]]
IIG_sheet = file[sheets[1]]
today = datetime.date.today()
yesterday = today-datetime.timedelta(days=1)
ITC_sheet["A1"].value = yesterday
ITC_sheet["E5"].value = float(airtel_in)
ITC_sheet["F5"].value = float(airtel_out)
ITC_sheet["E6"].value =  float(uk_in)
ITC_sheet["F6"].value =  float(uk_out)
IIG_sheet["E8"].value =  float(bsccl_in)
IIG_sheet["F8"].value =  float(bsccl_out)
ITC_sheet["E7"].value =  float(sg_SMW4_in)
ITC_sheet["F7"].value =  float(sg_SMW4_out)
ITC_sheet["E12"].value = float(fna_in)
ITC_sheet["F12"].value = float(fna_out)
ITC_sheet["E11"].value = float(akamai_in)
ITC_sheet["F11"].value = float(akamai_out)
ITC_sheet["E16"].value = float(singtel_in)
ITC_sheet["F16"].value = float(singtel_out)
ITC_sheet["E17"].value = float(TATA_in)
ITC_sheet["F17"].value = float(TATA_out)
ITC_sheet["E22"].value = float(digicon_in)
ITC_sheet["F22"].value = float(digicon_out)
ITC_sheet["L25"].value = float(Novotel_IPLC_in)
ITC_sheet["M25"].value = float(Novotel_IPLC_out)
ITC_sheet["L26"].value = float(Novotel_SG_SW_in)
ITC_sheet["M26"].value = float(Novotel_SG_SW_out)
ITC_sheet["E24"].value = float(UIL_SG_in)
ITC_sheet["F24"].value = float(UIL_SG_out)
ITC_sheet["E27"].value = float(level3_in)
ITC_sheet["F27"].value = float(level3_out)
ITC_sheet["E30"].value = float(nv_uk_in)
ITC_sheet["F30"].value = float(nv_uk_out)
ITC_sheet["E31"].value = float(dbl_in)
ITC_sheet["F31"].value = float(dbl_out)
ITC_sheet["L5"].value =  float(google_node_1_in)
ITC_sheet["M5"].value =  float(google_node_1_out)
ITC_sheet["L6"].value =  float(google_node_2_in)
ITC_sheet["M6"].value =  float(google_node_2_out)
ITC_sheet["L17"].value =  "=SUM(L5:L6)"
ITC_sheet["M17"].value =  "=SUM(M5:ML6)"

shutil.move("../NovoCom-BW-Report.xlsx", "/shared/Bandwidth Report/Old/NovoCom-BW-Report.xlsx")
file.save("../NovoCom-BW-Report.xlsx")
file.close()

# Start Of interCloud bandwidht report
print("Putting all the values in InterCloud bandwidth report.........")

raw_bandwidth_file =  openpyxl.load_workbook("IC_raw_bw.xlsx")
raw_sheet = None
raw_sheets = raw_bandwidth_file.sheetnames
for item in raw_sheets:
    raw_sheet = raw_bandwidth_file[item]
    IC_novocom_in = raw_sheet["B1"].value
    IC_novocom_out = raw_sheet["C1"].value
    BDIX_in = raw_sheet["B2"].value
    BDIX_out = raw_sheet["C2"].value

raw_bandwidth_file.close()

file =  openpyxl.load_workbook("../IC-BW Report.xlsx")
sheets = file.sheetnames
sheet = None

IC_sheet = file[sheets[0]]

today = datetime.date.today()
yesterday = today-datetime.timedelta(days=1)
IC_sheet["A1"].value = yesterday
IC_sheet["E5"].value = float(IC_novocom_in)
IC_sheet["F5"].value = float(IC_novocom_out)
IC_sheet["E6"].value = float(BDIX_in)
IC_sheet["F6"].value = float(BDIX_out)

shutil.move("../IC-BW Report.xlsx", "/shared/Bandwidth Report/Old/IC-BW Report.xlsx")
file.save("../IC-BW Report.xlsx")
file.close()
