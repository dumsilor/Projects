class excel_parcer:


    def __init__(self,file,cell_start,cell_end):
        self.file = file
        self.cell_start = cell_start
        self.cell_end = cell_end

    def make_list_client (self):
        from openpyxl import load_workbook
        import re
        #from openpyxl import load_workbook
        workbook = load_workbook(filename = self.file)
        client_list = []
        for worksheet in workbook:
            status =False
            for cell_range in worksheet[self.cell_start:self.cell_end]:
                if status == True:
                    break
                for cell in cell_range:
                    if cell.value == "Total":
                        status = True
                        row_start = cell.row
                        break
                    else:
                        client_list.append(cell.value)
            break
        workbook.close()
        next_cell = "C"+str(row_start+1)
        return client_list, next_cell



ent_client_list = excel_parcer("IC-BW Report.xlsx",'C11','C500')
lsp_start_cell = ent_client_list.make_list_client()[1]

lsp_client_list = excel_parcer("IC-BW Report.xlsx",lsp_start_cell,'C500')
ggc_start_cell = lsp_client_list.make_list_client()[1]

ggc_clie
