from openpyxl import load_workbook
import re
client_list = []
row_count = 10
wb = load_workbook(filename="IC-BW Report.xlsx")

lsp_start_cell = None
ggc_start_cell = None
#Loop for Enterprise client name
for ws in wb:
    status = False
    for sheet in ws['C11':'C500']:
        if status == True:
            break
        for cell in sheet:
            if cell.value == "Total":
                status = True
                lsp_start_cell =cell.row
                break
            else:
                client_list.append(cell.value)
            row_count += 1
    break
client_bandwidth = []
sum_string = r'\=[SUM]'

#Loop For Enterprise Client bandwidth
for ws in wb:
    status = False
    for sheet in ws['D11':'D500']:
        if status == True:
            break
        for cell in sheet:
            if str(cell.value).startswith("=SUM"):
                status = True
                break
            else:
                client_bandwidth.append(cell.value)
    break

#loop for enterprise remarks
end_row = "F"+str(row_count)
#print(end_row)
ent_client_remarks=[]

for ws in wb:
    status = False
    for sheet in ws['F11':end_row]:
        if status == True:
            break
        for cell in sheet:
            if cell.value == None:
                ent_client_remarks.append('')
            else:
                ent_client_remarks.append(cell.value)
            end_cell = cell.row
    break


#start of LSP client:
row_count = 10
client_count =0
lsp_start_cell_name= "C"+str(lsp_start_cell+1)
lsp_start_cell_bw= "D"+str(lsp_start_cell+1)
lsp_client_list = []
lsp_client_bandwidth = []




#Loop for LSP client name
for ws in wb:
    status = False
    for sheet in ws[lsp_start_cell_name:'C500']:
        if status == True:
            break
        for cell in sheet:
            if cell.value == "Total":
                status = True
                ggc_start_cell =cell.row
                break
            else:
                lsp_client_list.append(cell.value)
            row_count += 1
            client_count+=1


    break


sum_string = r'\=[SUM]'

#Loop For LSP Client bandwidth
for ws in wb:
    status = False
    for sheet in ws[lsp_start_cell_bw:'D500']:
        if status == True:
            break
        for cell in sheet:
            if str(cell.value).startswith("=SUM"):
                status = True
                break
            else:
                lsp_client_bandwidth.append(cell.value)


    break


#loop for lsp remarks
start_row = "F"+str(end_cell+2)
end_row = "F"+str(end_cell+client_count+2)
lsp_client_remarks=[]

for ws in wb:
    status = False
    for sheet in ws[start_row:end_row]:
        if status == True:
            break
        for cell in sheet:
            if cell.value == None:
                lsp_client_remarks.append('')
            else:
                lsp_client_remarks.append(cell.value)
    break

#print(start_row,end_row)

#start of GGC client:
client_count =0
ggc_start_cell_name= "C"+str(ggc_start_cell+1)
ggc_start_cell_bw= "D"+str(ggc_start_cell+1)
ggc_client_list = []
ggc_client_bandwidth = []
row_count = 10



#Loop for GGC client name
for ws in wb:
    status = False
    for sheet in ws[ggc_start_cell_name:'C500']:
        if status == True:
            break
        for cell in sheet:
            if cell.value == "Total":
                status = True
                break
            else:
                ggc_client_list.append(cell.value)
            row_count +=1
            client_count +=1

    break


sum_string = r'\=[SUM]'

#Loop For GGC Client bandwidth
for ws in wb:
    status = False
    for sheet in ws[ggc_start_cell_bw:'D500']:
        if status == True:
            break
        for cell in sheet:
            if str(cell.value).startswith("=SUM"):
                status = True
                break
            else:
                ggc_client_bandwidth.append(cell.value)
    break



#loop for ggc remarks
start_row = "F"+str(end_cell+2)
end_row = "F"+str(end_cell+client_count+2)
#print(end_row)
ggc_client_remarks=[]

for ws in wb:
    status = False
    for sheet in ws[start_row:end_row]:
        if status == True:
            break
        for cell in sheet:
            if cell.value == None:
                ggc_client_remarks.append('')
            else:
                ggc_client_remarks.append(cell.value)
    break


#Create table row for HTML table
#GGC client
ggc_client = None
x = 0
while x<len(ggc_client_list):
    temp_client= ggc_client_list[x]
    temp_bw = ggc_client_bandwidth[x]
    temp_remark = ggc_client_remarks[x]
    if ggc_client == None:
        ggc_client ="<tr>\n"+"<td>"+str(x+1)+"</td>"+"<td>"+temp_client +"</td>"+"<td>"+str(temp_bw) +"</td>"+"<td>"+temp_remark +"</td>\n"+"<tr>\n"
    else:
        ggc_client = ggc_client+"<tr>\n"+"<td>"+str(x+1)+"</td>"+"<td>"+temp_client +"</td>"+"<td>"+str(temp_bw) +"</td>"+"<td>"+temp_remark +"</td>\n"+"<tr>\n"
    x +=1

#enterprise Client:
ent_client = None
x = 0
while x<len(client_bandwidth):
    temp_client= client_list[x]
    temp_bw = client_bandwidth[x]
    temp_remark = ent_client_remarks[x]
    if ent_client == None:
        ent_client ="<tr>\n"+"<td>"+str(x+1)+"</td>"+"<td>"+temp_client +"</td>"+"<td>"+str(temp_bw) +"</td>"+"<td>"+temp_remark +"</td>\n"+"<tr>\n"
    else:
        ent_client = ent_client+"<tr>\n"+"<td>"+str(x+1)+"</td>"+"<td>"+temp_client +"</td>"+"<td>"+str(temp_bw) +"</td>"+"<td>"+temp_remark +"</td>\n"+"<tr>\n"
    x +=1

#lsp Client:
lsp_client = None
x = 0
while x<len(lsp_client_list):
    temp_client= lsp_client_list[x]
    temp_bw = lsp_client_bandwidth[x]
    temp_remark = lsp_client_remarks[x]
    if lsp_client == None:
        lsp_client ="<tr>\n"+"<td>"+str(x+1)+"</td>"+"<td>"+temp_client +"</td>"+"<td>"+str(temp_bw) +"</td>"+"<td>"+temp_remark +"</td>\n"+"<tr>\n"
    else:
        lsp_client = lsp_client+"<tr>\n"+"<td>"+str(x+1)+"</td>"+"<td>"+temp_client +"</td>"+"<td>"+str(temp_bw) +"</td>"+"<td>"+temp_remark +"</td>\n"+"<tr>\n"
    x +=1




#print(client)

#HTML Table generate

from datetime import date, timedelta

yesterday = date.today() - timedelta(days=1)
date = yesterday.strftime('%d/%m/%y')


filename = "table.html"
fhandle = open(filename, "w")
mail_body = """
<html>
    <head>
        <style>
            tr, td, th { border: 1px solid black;
                         text-align: center;
                         padding: 15px;
                          }
            #date {
            width: 10%;
            }
            .head {
            background-color: 
            }

        </style>
    </head>
    <body>
        <p> Dear Sir, <br>
        Kindly find the attached mail for InterCloud Bandwidth report:
        </p>
        <table>
            <thead>
                <tr>
                    <th id=date>"""+date+"""</th>
                    <th id=report_head colspan ="3"> InterCloud Bandwidth Report</th>
                </tr>
                <tr class=head>
                    <th colspan="4">Enterprise Client</th>
                </tr>
            </thead>
            <tbody>""" + ent_client + """</tbody>
            <thead>
                <tr class=head>
                    <th colspan="4">LSP Client</th>
                </tr>
            </thead>
            <tbody>""" + lsp_client + """</tbody>
            <thead>
                <tr class=head>
                    <th colspan="4">GGC Client</th>
                </tr>
            </thead>
            <tbody>""" + ggc_client + """</tbody>
        </table>
    </body>
</html>"""
fhandle.write(mail_body)
fhandle.close()
