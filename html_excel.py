from openpyxl import load_workbook
import re
client_list = []

wb = load_workbook(filename="IC-BW Report.xlsx")
#filename = "table.html"
#fhandle = open(filename, "w")
#fhandle.write("<html>")
#fhandle.write("<head>")
#fhandle.write("<style>")
#fhandle.write("tr, td { border: 1px solid black; }")
#fhandle.write("</style>")
#fhandle.write("</head>")
#fhandle.write("<body>")
#fhandle.write("<table>")

lsp_start_cell = None
ggc_start_cell = None
#Loop for Enterprise client name
for ws in wb:
    status = False
    for sheet in ws['C11':'C500']:
        if status == True:
            break
        for cell in sheet:
            if cell.value == "Total":
                status = True
                lsp_start_cell =cell.row
                break
            else:
                client_list.append(cell.value)


    break
client_bandwidth = []
sum_string = r'\=[SUM]'

#Loop For Enterprise Client bandwidth
for ws in wb:
    status = False
    for sheet in ws['D11':'D500']:
        if status == True:
            break
        for cell in sheet:
            if str(cell.value).startswith("=SUM"):
                status = True
                break
            else:
                client_bandwidth.append(cell.value)

    break



lsp_start_cell_name= "C"+str(lsp_start_cell+1)
lsp_start_cell_bw= "D"+str(lsp_start_cell+1)
lsp_client_list = []
lsp_client_bandwidth = []




#Loop for LSP client name
for ws in wb:
    status = False
    for sheet in ws[lsp_start_cell_name:'C500']:
        if status == True:
            break
        for cell in sheet:
            if cell.value == "Total":
                status = True
                ggc_start_cell =cell.row
                break
            else:
                lsp_client_list.append(cell.value)


    break


client_bandwidth = []
sum_string = r'\=[SUM]'

#Loop For LSP Client bandwidth
for ws in wb:
    status = False
    for sheet in ws[lsp_start_cell_bw:'D500']:
        if status == True:
            break
        for cell in sheet:
            if str(cell.value).startswith("=SUM"):
                status = True
                break
            else:
                lsp_client_bandwidth.append(cell.value)


    break






ggc_start_cell_name= "C"+str(ggc_start_cell+1)
ggc_start_cell_bw= "D"+str(ggc_start_cell+1)
ggc_client_list = []
ggc_client_bandwidth = []




#Loop for GGC client name
for ws in wb:
    status = False
    for sheet in ws[ggc_start_cell_name:'C500']:
        if status == True:
            break
        for cell in sheet:
            if cell.value == "Total":
                status = True
                break
            else:
                ggc_client_list.append(cell.value)


    break


client_bandwidth = []
sum_string = r'\=[SUM]'

#Loop For GGC Client bandwidth
for ws in wb:
    status = False
    for sheet in ws[ggc_start_cell_bw:'D500']:
        if status == True:
            break
        for cell in sheet:
            if str(cell.value).startswith("=SUM"):
                status = True
                break
            else:
                ggc_client_bandwidth.append(cell.value)


    break


                #fhandle.write("<td>" + str(cell.value)+ "</td>")

#        fhandle.write("</tr>")
#fhandle.write("</table>")
#fhandle.write("</body>")
#fhandle.write("</html>")
#fhandle.close()
#sheet_ranges = wb['IIG']
#for element in sheet_ranges:
#    for cell in element:
#        if cell.value:
#            print(cell.value)
