########################################################################

#Processing Excel using VBA

########################################################################
from openpyxl import load_workbook
import os

class excel_parcer:

    def __init__(self,file,cell_start,cell_end,sheet_number):
        self.file = file
        self.cell_start = cell_start
        self.cell_end = cell_end
        os.chdir("/shared/Bandwidth Report")
        workbook = load_workbook(filename = self.file)
        self.workbook = workbook
        sheets = workbook.sheetnames
        sheet = sheets[sheet_number]
        self.sheet = sheet

    def make_list_client (self):
        workbook = self.workbook
        sheet = self.sheet
        worksheet = workbook[sheet]
        client_list = []
        start = 'C'+self.cell_start
        end ='C'+self.cell_end
        status =False
        for cell_range in worksheet[start:end]:
            if status == True:
                break
            for cell in cell_range:
                if cell.value == "Total":
                    status = True
                    row_start = cell.row
                    break
                else:
                    client_list.append(cell.value)
        row_count = len(client_list)
        next_cell = str(row_start+1)
        workbook.close()
        return client_list, next_cell


    def make_list_bw (self):
        workbook = self.workbook
        sheet = self.sheet
        worksheet = workbook[sheet]
        bw_list = []
        start = 'D'+self.cell_start
        end ='D'+self.cell_end
        status =False
        for cell_range in worksheet[start:end]:
            if status == True:
                break
            for cell in cell_range:
                if str(cell.value).startswith("=SUM"):
                    status = True
                    break
                else:
                    bw_list.append(cell.value)
            workbook.close()
        return bw_list

    def make_list_remarks (self,list_len,index,column):
        workbook = self.workbook
        sheet = self.sheet
        worksheet = workbook[sheet]
        remarks_list = []
        start = column+self.cell_start
        end = column+str(list_len+index) #index = client list lenght + previous client type start row
        #print("start: ",start)
        #print('end: ',end)
        #print("index:" ,index)
        #print("list len: ",list_len)
        status =False
        for cell_range in worksheet[start:end]:
            if status == True:
                break
            for cell in cell_range:
                if cell.value == None:
                    remarks_list.append('')
                else:
                    remarks_list.append(cell.value)
            workbook.close()
        return remarks_list

    def make_html (self,count,client_list,bw_list,remarks_list):
        x=0
        client = None
        self.count = count
        #print(remarks_list)
        while x < int(self.count):
            temp_client = client_list[x]
            temp_bw = bw_list[x]
            temp_remark = remarks_list[x]
            if client == None:
                client ="<tr class = objects>\n"+"<td>"+str(x+1)+"</td>"+"<td>"+temp_client +"</td>"+"<td>"+str(temp_bw) +"</td>"+"<td colspan='5'>"+temp_remark +"</td>\n"+"</tr>\n"
            else:
                client = client+"<tr class = objects>\n"+"<td>"+str(x+1)+"</td>"+"<td>"+temp_client +"</td>"+"<td>"+str(temp_bw) +"</td>"+"<td colspan='5'>"+temp_remark +"</td>\n"+"</tr>\n"
            x +=1
        return client

    def upstream (self):

         upstream_in =  []
         upstream_out = []
         upstream_name = []
         total_bw = []
         columns = ["C","D","E","F"]
         #from openpyxl import load_workbook
         workbook = self.workbook
         sheet = self.sheet
         ic_bw_sheet = workbook[sheet]
         for column in columns:
            start = column + self.cell_start
            end = column + self.cell_end
            for cell_range in ic_bw_sheet[start:end]:
                for cell in cell_range:
                    if column == "E":
                        upstream_in.append(cell.value)
                    elif column == "F":
                        upstream_out.append(cell.value)
                    elif column == "C":
                        upstream_name.append(cell.value)
                    elif column == "D":
                        total_bw.append(cell.value)
         return upstream_in,upstream_out, upstream_name, total_bw

    def percentage (self,bw,total):
        percent = (bw/total)*100
        return round(percent,2)

    def upstream_html(self,tuple):
        upstream_name = tuple[2]
        upstream_total = tuple[3]
        upstream_in = tuple[0]
        upstream_out =tuple[1]


        serial = 1
        index = 0
        upstream = None
        for name in upstream_name:

            upstream_in_util = self.percentage(upstream_in[index], upstream_total[index])
            upstream_out_util = self.percentage(upstream_out[index], upstream_total[index])
            if upstream == None:
                upstream = "<tr class = objects>\n" + "<td>" + str(serial) + "</td>" + "<td>" + str(upstream_name[index]) + "</td>" + "<td>" + str(upstream_total[index]) + "</td>" + "<td>" + str(upstream_in[index]) + "</td>" +"<td>"+ str(upstream_out[index]) + "</td>" + "<td>" + str(upstream_in_util) + "</td>" + "<td>" + str(upstream_out_util) + "</td>\n" +"<td></td>"+ "</tr>\n"
            else:
                upstream = upstream+"<tr class = objects>\n" + "<td>" + str(serial) + "</td>" + "<td>" + str(upstream_name[index]) + "</td>" + "<td>" + str(upstream_total[index]) + "</td>" + "<td>" + str(upstream_in[index]) + "</td>" +"<td>"+ str(upstream_out[index]) + "</td>" + "<td>" + str(upstream_in_util) + "</td>" + "<td>" + str(upstream_out_util) + "</td>\n" +"<td></td>"+ "</tr>\n"
            serial += 1
            index +=1
        return upstream



    def get_cell_value(self,cell_add):
        workbook = self.workbook
        sheet = self.sheet
        worksheet = workbook[sheet]
        cell = worksheet[cell_add]
        cell_value = cell.value
        return cell_value




########################### end of Class #########################

def total_bw(bw_list):
    total = 0
    for bw in bw_list:
        if bw == None:
            continue
        else:
            total = total+bw
    return total



########################################################################

#enterprise Client

########################################################################
filename ="IC-BW-Report.xlsx"

ent_client_start_row = 11
ent_client = excel_parcer(filename,str(ent_client_start_row),'500',0)
lsp_start_cell = ent_client.make_list_client()[1]
ent_client_list = ent_client.make_list_client()[0]
ent_client_bw = ent_client.make_list_bw()
ent_total = total_bw (ent_client_bw)
ent_client_remarks = ent_client.make_list_remarks(len(ent_client_list)+10,0,"F")
ent_count=len(ent_client_list)
ent_html = ent_client.make_html(ent_count,ent_client_list,ent_client_bw,ent_client_remarks)

########################################################################

#LSP Client

########################################################################

lsp_client = excel_parcer(filename,lsp_start_cell,'500',0)
ggc_start_cell = lsp_client.make_list_client()[1]
lsp_client_list = lsp_client.make_list_client()[0]
lsp_client_bw = lsp_client.make_list_bw()
lsp_total = total_bw (lsp_client_bw)
lsp_client_remarks = lsp_client.make_list_remarks(len(lsp_client_list),ent_count+ent_client_start_row,"F") #Enterprise start row number = 11
#print(lsp_client_remarks)
lsp_count = len(lsp_client_list)
lsp_html = lsp_client.make_html(lsp_count,lsp_client_list,lsp_client_bw,lsp_client_remarks)

########################################################################

#GGC Client

########################################################################

ggc_client = excel_parcer(filename,ggc_start_cell,'500',0)
ggc_client_list = ggc_client.make_list_client()[0]
ggc_client_bw = ggc_client.make_list_bw()
ggc_total = total_bw (ggc_client_bw)
ggc_client_remarks = ggc_client.make_list_remarks(len(ggc_client_list), lsp_count+int(lsp_start_cell),"F")
ggc_count = len(ggc_client_list)
ggc_html = lsp_client.make_html(ggc_count,ggc_client_list,ggc_client_bw,ggc_client_remarks)


########################################################################

#Upstream Bandwidth

########################################################################

upstream = excel_parcer(filename,"5",'6',0)
upstream_list = upstream.upstream()
up_html = upstream.upstream_html(upstream_list)





class send_mail:
    def __init__(self,sender, receiver_list,subject,caption,html_body,filename):
        self.filename = filename
        self.sender = sender
        self.receiver_list = receiver_list
        self.subject = subject
        self.caption =caption
        self.html_body = html_body

        import email, smtplib, ssl

        from email import encoders
        from email.mime.base import MIMEBase
        from email.mime.multipart import MIMEMultipart
        from email.mime.text import MIMEText

        password = "ipcore#@!"

        message = MIMEMultipart()
        message["Subject"] = subject
        message["From"] = sender
        message["To"] = ','.join(receiver_list)
        message["Cc"] = ','.join(caption)


        # Turn these into plain/html MIMEText objects
        #part1 = MIMEText(text, "plain")
        part2 = MIMEText(html_body, "html")

        # Add HTML/plain-text parts to MIMEMultipart message
        # The email client will try to render the last part first
        #message.attach(part1)
        message.attach(part2)
        if filename == "":
            text = message.as_string()
            context = ssl.create_default_context()
            with smtplib.SMTP_SSL("mta.brilliant.com.bd", 465, context=context) as server:
                server.login(sender, password)
                server.sendmail(
                    sender, receiver_list + caption ,text
                )
        else:
            with open(filename, "rb") as attachment:
                # Add file as application/octet-stream
                # Email client can usually download this automatically as attachment
                part = MIMEBase("application", "octet-stream")
                part.set_payload(attachment.read())

                # Encode file in ASCII characters to send by email
            encoders.encode_base64(part)

                # Add header as key/value pair to attachment part
            part.add_header(
            "Content-Disposition",
            f"attachment; filename= {filename}",
            )

                # Add attachment to message and convert message to string
            message.attach(part)
            text = message.as_string()


                # Create secure connection with server and send email
            context = ssl.create_default_context()
            with smtplib.SMTP_SSL("mta.brilliant.com.bd", 465, context=context) as server:
                server.login(sender, password)
                server.sendmail(
                sender, receiver_list + caption ,text
                )










########################################################################

#Creating HTML

########################################################################


from datetime import date, timedelta

yesterday = date.today() - timedelta(days=1)
date = yesterday.strftime('%d/%m/%y')

name = input("Please Enter Shift Engineer Name: ")


mail_body = """
<html>
    <head>
        <style>
            table {
            padding: 0 20px 0 20px;
            }
            tr, td, th { border: 1px solid black;
                         text-align: center;
                         padding: 15px;
                          }
            #date {
            width: 10%;
            }


            .table_head {
            background: #a3a375;
            color: #800000;
            font-size: 120%;
            }

            #report_head, #date {
            font-size:130%;
            color: #800000;
            background : #999966
            }
            #date {
            font-size:100%;
            }
            .objects {
            background : #d6d6c2
            }
            .head {
            background-color:#c1c1a4;
            }
            .total {
            font-weight: bold;
            background : #d6d6c2
            }
        </style>
    </head>
    <body>
        <p> Dear Sir, <br>
        <br>
        Kindly find the attached mail for InterCloud Bandwidth report:
        </p>
        <table>
            <thead>
                <tr>
                    <th id=date>"""+date+"""</th>
                    <th id=report_head colspan ="7"> InterCloud Bandwidth Report</th>
                </tr>

                <tr>
                    <th colspan="8" class ="Table_head">Upstream</th>
                </tr>

                <tr class=head>
                    <th>Sr. Number</th>
                    <th>Upstream Name</th>
                    <th>Total Capacity(Mbps)</th>
                    <th>Highest Bandwidth IN(Mbps)</th>
                    <th>Highest Bandwidth Out(Mbps)</th>
                    <th>Utilization IN (%) </th>
                    <th>Utilization OUT (%) </th>
                    <th>Remarks</th>
                </tr>

                """+up_html+"""

                <tr>
                    <th colspan="8" class ="Table_head">Enterprise Client</th>
                </tr>
                <tr class=head>
                    <th>Sr. Number</th>
                    <th>Name</th>
                    <th>Bandwidth</th>
                    <th colspan='5'>Remarks</th>
                </tr>
            </thead>
            <tbody>""" + ent_html + """
            <tr class = 'total'>
                <td colspan ='2'>Total: </td>
                <td>"""+ str(ent_total) + """</td>
                <td colspan ='5'> </td>
            </tr>
            </tbody>
            <thead>
                <tr>
                    <th colspan="8" class ="Table_head">LSP Client</th>
                </tr>
                <tr class=head>
                    <th>Sr. Number</th>
                    <th>Name</th>
                    <th>Bandwidth</th>
                    <th colspan='5'>Remarks</th>
                </tr>
            </thead>
            <tbody>""" + lsp_html + """
            <tr class = 'total'>
                <td colspan ='2'>Total: </td>
                <td>"""+ str(lsp_total) + """</td>
                <td colspan ='5'> </td>
            </tr>

            </tbody>
            <thead>
                <tr>
                    <th colspan="8" class ="Table_head">GGC Client</th>
                </tr>
                <tr class=head>
                    <th>Sr. Number</th>
                    <th>Name</th>
                    <th>Bandwidth</th>
                    <th colspan='5'>Remarks</th>
                </tr>
            </thead>
            <tbody>""" + ggc_html + """

            <tr class = 'total'>
                <td colspan ='2'>Total: </td>
                <td>"""+ str(ggc_total) + """</td>
                <td colspan ='5'> </td>
            </tr>

            </tbody>
        </table>
        <br>
        <br>
        <p>Regards,</p>
        <br>
        <p>""" + name + """</p>
        <p><b>Core Network || AS58923</b></p>
        <p><b>Internet Service Provider (ISP) & IP Telephony Service Provider (IPTSP)</b></p>
        <p>House: Ga-30/G, Pragati Sarani, Shahjadpur, Gulshan-2, Dhaka – 1212, Bangladesh</p>
        <p><b>phone:</b>  +8809638383838; Ext: 8426, 8427</p>
    </body>
</html>"""






########################################################################

#Creating Tunnel and Sending Mail For Intercloud

########################################################################
print ("Sending IntrtCloud Report")
sender_email = "core@brilliant.com.bd"
receiver_email = ["revenue.assurance@intercloud.com.bd","info@intercloud.com.bd","bivabory.nayana@intercloud.com.bd","s.hossain@intercloud.com.bd"]
caption = ["core@brilliant.com.bd","noc@brilliant.com.bd"]
subject = "Bandwidth Utilization Report of Intercloud"

send_mail(sender_email,receiver_email,subject,caption,mail_body,filename)


print("InterCloud Report sent Successfully!")


########################################################################

#Bandwidth Report For NovoCom ITC

########################################################################
filename ="NovoCom-BW-Report.xlsx"
iig_sheet_number = 1
itc_sheet_number = 0

###############   ITC Upstream   #######################################

#ITC upstream started from row 5 and ended in row 7 need to change if anything changes

itc_upstream = excel_parcer(filename,"5",'7',itc_sheet_number) #Filname to open, start row number, end row number, Worksheetnumber
itc_upstream_list = itc_upstream.upstream()

itc_upstream_html = itc_upstream.upstream_html(itc_upstream_list)


###############   ITC Downstream  ##################################

#ITC downstream started from row 9 and ended in row 9 need to change if anything changes


itc_downstream = excel_parcer(filename,"9",'9',itc_sheet_number) #Filname to open, start row number, end row number, Worksheetnumber
itc_downstream_list = itc_downstream.upstream() # output structure ([inbound,outbound,provider_name,Total])

itc_downstream_list[0][0] = itc_upstream_list[0][0] + itc_upstream_list[0][1]
itc_downstream_list[1][0] = itc_upstream_list[1][0] + itc_upstream_list[1][1]

itc_downstream_list[1][0] = round(itc_downstream_list[1][0],2)

itc_downstream_html = itc_downstream.upstream_html(itc_downstream_list)


####################### Hosted Service #######################

hosted_service = excel_parcer(filename,'11','13',itc_sheet_number)
hosted_service_list = hosted_service.upstream()
ggc_node_1_val_in = hosted_service.get_cell_value("L5")
ggc_node_2_val_in = hosted_service.get_cell_value("L6")
ggc_node_1_val_out = hosted_service.get_cell_value("M5")
ggc_node_2_val_out = hosted_service.get_cell_value("M6")

ggc_in = ggc_node_1_val_in + ggc_node_2_val_in
ggc_out =  ggc_node_1_val_out + ggc_node_2_val_out

hosted_service_list[0][2] = ggc_in
hosted_service_list[1][2] = ggc_out

hosted_service_html = hosted_service.upstream_html(hosted_service_list)



####################### SG upstream #######################


sg_upstream = excel_parcer(filename,"16",'19',itc_sheet_number) #Filname to open, start row number, end row number, Worksheetnumber
sg_upstream_list = sg_upstream.upstream()

sg_upstream_html = sg_upstream.upstream_html(sg_upstream_list)


####################### SG Downstream #######################

sg_downstream = excel_parcer(filename,"21",'24',itc_sheet_number) #Filname to open, start row number, end row number, Worksheetnumber
sg_downstream_list = sg_downstream.upstream()

sg_downstream_list[0][0] = itc_upstream_list[0][2]
sg_downstream_list[1][0] = itc_upstream_list[1][2]

novotel_iplc_in = hosted_service.get_cell_value("L25")
novotel_iplc_out = hosted_service.get_cell_value("M25")
novotel_sw_in = hosted_service.get_cell_value("L26")
novotel_sw_out = hosted_service.get_cell_value("M26")

novotel_sg_in = novotel_iplc_in + novotel_sw_in
novotel_sg_out = novotel_iplc_out + novotel_sw_out

sg_downstream_list[0][1] = itc_upstream_list[0][2]
sg_downstream_list[1][1] = itc_upstream_list[1][2]

sg_downstream_list[0][2] = novotel_sg_in
sg_downstream_list[1][2] = novotel_sg_out

sg_downstream_html = sg_downstream.upstream_html(sg_downstream_list)


####################### UK upstream #######################


uk_upstream = excel_parcer(filename,"27",'27',itc_sheet_number) #Filname to open, start row number, end row number, Worksheetnumber
uk_upstream_list = uk_upstream.upstream()

uk_upstream_html = uk_upstream.upstream_html(uk_upstream_list)


####################### UK Downstream #######################


uk_downstream = excel_parcer(filename,"29",'31',itc_sheet_number) #Filname to open, start row number, end row number, Worksheetnumber
uk_downstream_list = uk_downstream.upstream()

uk_downstream_list[0][0] = itc_upstream_list[0][1]
uk_downstream_list[1][0] = itc_upstream_list[1][1]

uk_downstream_html = uk_downstream.upstream_html(uk_downstream_list)


########################################################################

#Bandwidth Report For NovoCom IIG Client

########################################################################



client_start_row = 13
iig_client = excel_parcer(filename,str(client_start_row),'500',iig_sheet_number)
iig_client_list = iig_client.make_list_client()[0]
intercloud_pos = iig_client_list.index("InterCloud")
iig_client_bw = iig_client.make_list_bw()
iig_total = total_bw (iig_client_bw) - iig_client_bw[intercloud_pos]
iig_client_remarks = iig_client.make_list_remarks(len(iig_client_list)+10,2,"E")
iig_count=len(iig_client_list)
iig_client_html = iig_client.make_html(iig_count,iig_client_list,iig_client_bw,iig_client_remarks)


########################################################################

#Bandwidth Report For NovoCom IIG Upstream

########################################################################

iig_upstream = excel_parcer(filename,"7",'9',iig_sheet_number) #Filname to open, start row number, end row number, Worksheetnumber
iig_upstream_list = iig_upstream.upstream()

iig_upstream_list[0][0]=itc_downstream_list[0][0]
iig_upstream_list[1][0]=itc_downstream_list[1][0]

iig_upstream_list[0][2]=itc_upstream_list[0][2]
iig_upstream_list[1][2]=itc_upstream_list[1][2]

iig_upstream_html = iig_upstream.upstream_html(upstream_list)

########################################################################

#Creating HTML for NovoCom IIG

########################################################################


iig_body = """
<html>
    <head>
        <style>
            table {
            padding: 0 20px 0 20px;
            }
            tr, td, th { border: 1px solid black;
                         text-align: center;
                         padding: 15px;
                          }
            #date {
            width: 10%;
            }


            .table_head {
            background: #a3a375;
            color: #800000;
            font-size: 120%;
            }

            #report_head, #date {
            font-size:130%;
            color: #800000;
            background : #999966
            }
            #date {
            font-size:100%;
            }
            .objects {
            background : #d6d6c2
            }
            .head {
            background-color:#c1c1a4;
            }
            .total {
            font-weight: bold;
            background : #d6d6c2
            }
        </style>
    </head>
    <body>
        <p> Dear Sir, <br>
        <br>
        Kindly find the attached mail for NovoCom IIG Bandwidth report:
        </p>
        <table>
            <thead>
                <tr>
                    <th id=date>"""+date+"""</th>
                    <th id=report_head colspan ="7"> NovoCom IIG Bandwidth Report</th>
                </tr>

                <tr>
                    <th colspan="8" class ="Table_head">Upstream</th>
                </tr>

                <tr class=head>
                    <th>Sr. Number</th>
                    <th>Upstream Name</th>
                    <th>Total Capacity(Mbps)</th>
                    <th>Highest Bandwidth IN(Mbps)</th>
                    <th>Highest Bandwidth Out(Mbps)</th>
                    <th>Utilization IN (%) </th>
                    <th>Utilization OUT (%) </th>
                    <th>Remarks</th>
                </tr>

                """+iig_upstream_html+"""

                <tr>
                    <th colspan="8" class ="Table_head">Client</th>
                </tr>
                <tr class=head>
                    <th>Sr. Number</th>
                    <th>Name</th>
                    <th>Bandwidth</th>
                    <th colspan='5'>Remarks</th>
                </tr>
            </thead>
            <tbody>""" + iig_client_html + """
            <tr class = 'total'>
                <td colspan ='2'>Total [excluding InterCloud] : </td>
                <td>"""+ str(iig_total) + """</td>
                <td colspan ='5'> </td>
            </tr>
            </tbody>
        </table>
        <br>
        <br>
        <p>Regards,</p>
        <br>
        <p>""" + name + """</p>
        <p><b>IIG NOC || AS58923</b></p>
        <p><b>Internet Service Provider (ISP) & IP Telephony Service Provider (IPTSP)</b></p>
        <p>House: Ga-30/G, Pragati Sarani, Shahjadpur, Gulshan-2, Dhaka – 1212, Bangladesh</p>
        <p><b>phone:</b>  +8809638383838; Ext: 8426, 8427</p>
    </body>
</html>"""

#fname = "table.html"
#fhandle = open(fname, "w")

########################################################################

#Creating HTML for NovoCom ITC

########################################################################


itc_body = """
<html>
    <head>
        <style>
            table {
            padding: 0 20px 0 20px;
            }
            tr, td, th { border: 1px solid black;
                         text-align: center;
                         padding: 15px;
                          }
            #date {
            width: 10%;
            }


            .table_head {
            background: #a3a375;
            color: #800000;
            font-size: 120%;
            }

            #report_head, #date {
            font-size:130%;
            color: #800000;
            background : #999966
            }
            #date {
            font-size:100%;
            }
            .objects {
            background : #d6d6c2
            }
            .head {
            background-color:#c1c1a4;
            }
            .total {
            font-weight: bold;
            background : #d6d6c2
            }
        </style>
    </head>
    <body>
        <p> Dear Sir, <br>
        <br>
        Kindly find the attached mail for NovoCom ITC Bandwidth report:
        </p>
        <table>
            <thead>
                <tr>
                    <th id=date>"""+date+"""</th>
                    <th id=report_head colspan ="7"> NovoCom Limited - ITC</th>
                </tr>

                <tr>
                    <th colspan="8" class ="Table_head">Upstream</th>
                </tr>

                <tr class=head>
                    <th>Sr. Number</th>
                    <th>Upstream Name</th>
                    <th>Total Capacity(Mbps)</th>
                    <th>Highest Bandwidth IN(Mbps)</th>
                    <th>Highest Bandwidth Out(Mbps)</th>
                    <th>Utilization IN (%) </th>
                    <th>Utilization OUT (%) </th>
                    <th>Remarks</th>
                </tr>
                """+itc_upstream_html+ """

                <tr>
                    <th colspan="8" class ="Table_head">Downstream</th>
                </tr>

                <tr class=head>
                    <th>Sr. Number</th>
                    <th>Upstream Name</th>
                    <th>Total Capacity(Mbps)</th>
                    <th>Highest Bandwidth IN(Mbps)</th>
                    <th>Highest Bandwidth Out(Mbps)</th>
                    <th>Utilization IN (%) </th>
                    <th>Utilization OUT (%) </th>
                    <th>Remarks</th>
                </tr>
                """+itc_downstream_html +"""

                <tr>
                    <th id=report_head colspan ="8"> NovoCom Limited - Singapore</th>
                </tr>

                <tr>
                    <th colspan="8" class ="Table_head">Upstream</th>
                </tr>

                <tr class=head>
                    <th>Sr. Number</th>
                    <th>Upstream Name</th>
                    <th>Total Capacity(Mbps)</th>
                    <th>Highest Bandwidth IN(Mbps)</th>
                    <th>Highest Bandwidth Out(Mbps)</th>
                    <th>Utilization IN (%) </th>
                    <th>Utilization OUT (%) </th>
                    <th>Remarks</th>
                </tr>
                """+sg_upstream_html+ """

                <tr>
                    <th colspan="8" class ="Table_head">Downstream</th>
                </tr>

                <tr class=head>
                    <th>Sr. Number</th>
                    <th>Upstream Name</th>
                    <th>Total Capacity(Mbps)</th>
                    <th>Highest Bandwidth IN(Mbps)</th>
                    <th>Highest Bandwidth Out(Mbps)</th>
                    <th>Utilization IN (%) </th>
                    <th>Utilization OUT (%) </th>
                    <th>Remarks</th>
                </tr>
                """+sg_downstream_html +"""

                <tr>
                    <th id=report_head colspan ="8"> NovoCom Limited - UK</th>
                </tr>

                <tr>
                    <th colspan="8" class ="Table_head">Upstream</th>
                </tr>

                <tr class=head>
                    <th>Sr. Number</th>
                    <th>Upstream Name</th>
                    <th>Total Capacity(Mbps)</th>
                    <th>Highest Bandwidth IN(Mbps)</th>
                    <th>Highest Bandwidth Out(Mbps)</th>
                    <th>Utilization IN (%) </th>
                    <th>Utilization OUT (%) </th>
                    <th>Remarks</th>
                </tr>
                """+uk_upstream_html+ """

                <tr>
                    <th colspan="8" class ="Table_head">Downstream</th>
                </tr>

                <tr class=head>
                    <th>Sr. Number</th>
                    <th>Upstream Name</th>
                    <th>Total Capacity(Mbps)</th>
                    <th>Highest Bandwidth IN(Mbps)</th>
                    <th>Highest Bandwidth Out(Mbps)</th>
                    <th>Utilization IN (%) </th>
                    <th>Utilization OUT (%) </th>
                    <th>Remarks</th>
                </tr>
                """+uk_downstream_html +"""
                </tbody>
            </table>
            <br>
            <br>
            <p>Regards,</p>
            <br>
            <p>""" + name + """</p>
            <p><b>IP Support NOC || AS132267</b></p>
            <p><b>Internet Service Provider (ISP) & IP Telephony Service Provider (IPTSP)</b></p>
            <p>House: Ga-30/G, Pragati Sarani, Shahjadpur, Gulshan-2, Dhaka – 1212, Bangladesh</p>
            <p><b>phone:</b>  +8809638383838; Ext: 8426, 8427</p>
        </body>
    </html>"""
#print(mail_body)
#fhandle.write(itc_body)
#fhandle.close()

########################################################################

#Send Mail for NovoCom IIG

########################################################################
print ("Sending IIG Report")

sender_email = "noc.iig@novocom-bd.com"
receiver_email = ["info@novocom-bd.com","bivabory@novocom-bd.com","s.hossain@novocom-bd.com","revenue.assurance@novocom-bd.com"]
caption = [sender_email]
subject = "Bandwidth Utilization Report of NovoCom IIG"

send_mail(sender_email,receiver_email,subject,caption,iig_body,filename)
print("IIG Report sent Successfully!")

########################################################################

#Send Mail for NovoCom ITC

########################################################################
print ("Sending ITC Report")

sender_email = "ipsupport@novocom-bd.com"
["info@novocom-bd.com","bivabory@novocom-bd.com","s.hossain@novocom-bd.com","revenue.assurance@novocom-bd.com"]
caption = [sender_email]
subject = "Bandwidth Utilization Report of NovoCom ITC"

send_mail(sender_email,receiver_email,subject,caption,itc_body,filename)


print("ITC Report sent Successfully!")

########################################################################

#Send Mail for NovoTel

########################################################################
filename = ""
nv_body = """'
<html>
    <body>
        <p>Dear Team, </p>
        <p> Kindly find the Internet Bandwidth utilization by NovoTel in Singapore and UK on """+date+""" as follows- </p>
        <p>Singapore Internet used by NovoTel = """+ str(sg_downstream_list[1][2])+ """ Mbps</p>
        <p>UK Internet used by NovoTel = """+ str(uk_downstream_list[1][1])+ """ Mbps</p>
        <p>Please note that the amount shared above is 80% accurate.</p>
        <br>
        <p>Regards, </p>
        <p>""" +name+"""</p>
        <br>
        <p> IP NOC</p>
        <p> NovoTel  Limited</p>
        <p>House No. Ga-30/G, Pragati Sarani, Shahjadpur, Gulshan-2, Dhaka-1212, Bangladesh.</p>
        <p> 24X7 Hotline:+8801787681191 || ipsupport@novotel-bd.com || www.novotel-bd.com</p>
    </body>
</html>
"""

print ("Sending Novotel Report")

sender_email = "ipsupport@novocom-bd.com"
receiver_email = ["noc@novotel-bd.com"]
caption = [sender_email, "ipsupport@novotel-bd.com"]
subject = "NovoTel Internet BW Utilization || " + date

send_mail(sender_email,receiver_email,subject,caption,nv_body,filename)
print("Novotel Report sent Successfully!")
