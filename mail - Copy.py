########################################################################

#Processing Excel using VBA

########################################################################

class excel_parcer:

    def __init__(self,file,cell_start,cell_end):
        self.file = file
        self.cell_start = cell_start
        self.cell_end = cell_end

    def make_list_client (self):
        from openpyxl import load_workbook
        import re
        #from openpyxl import load_workbook
        workbook = load_workbook(filename = self.file)
        client_list = []
        start = 'C'+self.cell_start
        end ='C'+self.cell_end
        for worksheet in workbook:
            status =False
            for cell_range in worksheet[start:end]:
                if status == True:
                    break
                for cell in cell_range:
                    if cell.value == "Total":
                        status = True
                        row_start = cell.row
                        break
                    else:
                        client_list.append(cell.value)
            break
        row_count = len(client_list)
        next_cell = str(row_start+1)
        workbook.close()
        return client_list, next_cell


    def make_list_bw (self):
        from openpyxl import load_workbook
        import re
        #from openpyxl import load_workbook
        workbook = load_workbook(filename = self.file)
        bw_list = []
        start = 'D'+self.cell_start
        end ='D'+self.cell_end

        for worksheet in workbook:
            status =False
            for cell_range in worksheet[start:end]:
                if status == True:
                    break
                for cell in cell_range:
                    if str(cell.value).startswith("=SUM"):
                        status = True
                        break
                    else:
                        bw_list.append(cell.value)
            break
            workbook.close()
        return bw_list

    def make_list_remarks (self,list_len,index):
        from openpyxl import load_workbook
        import re
        #from openpyxl import load_workbook
        workbook = load_workbook(filename = self.file)
        remarks_list = []
        start = 'F'+self.cell_start
        end = 'F'+str(list_len+index) #index = client list lenght + previous client type start row
        #print("start: ",start)
        #print('end: ',end)
        #print("index:" ,index)
        #print("list len: ",list_len)
        for worksheet in workbook:
            status =False
            for cell_range in worksheet[start:end]:
                if status == True:
                    break
                for cell in cell_range:
                    if cell.value == None:
                        remarks_list.append('')
                    else:
                        remarks_list.append(cell.value)
            break
            workbook.close()
        return remarks_list

    def make_html (self,count,client_list,bw_list,remarks_list):
        x=0
        client = None
        self.count = count
        #print(type(count))
        while x < int(self.count):
            temp_client = client_list[x]
            temp_bw = bw_list[x]
            temp_remark = remarks_list[x]
            if client == None:
                client ="<tr class = objects>\n"+"<td>"+str(x+1)+"</td>"+"<td>"+temp_client +"</td>"+"<td>"+str(temp_bw) +"</td>"+"<td colspan='5'>"+temp_remark +"</td>\n"+"</tr>\n"
            else:
                client = client+"<tr class = objects>\n"+"<td>"+str(x+1)+"</td>"+"<td>"+temp_client +"</td>"+"<td>"+str(temp_bw) +"</td>"+"<td colspan='5'>"+temp_remark +"</td>\n"+"</tr>\n"
            x +=1
        return client

    def upstream (self):
         from openpyxl import load_workbook
         import re
         upstream_in =  []
         upstream_out = []
         upstream_name = []
         total_bw = []
         columns = ["C","D","E","F"]
         #from openpyxl import load_workbook
         workbook = load_workbook(filename = self.file)
         sheets = workbook.sheetnames
         sheet = sheets[0]
         ic_bw_sheet = workbook[sheet]
         for column in columns:
            start = column + self.cell_start
            end = column + self.cell_end
            for cell_range in ic_bw_sheet[start:end]:
                for cell in cell_range:
                    if column == "E":
                        upstream_in.append(cell.value)
                    elif column == "F":
                        upstream_out.append(cell.value)
                    elif column == "C":
                        upstream_name.append(cell.value)
                    elif column == "D":
                        total_bw.append(cell.value)
         return upstream_in,upstream_out, upstream_name, total_bw

    def percentage (self,bw,total):
        percent = (bw/total)*100
        return round(percent,2)

    def upstream_html(self,tuple):
        upstream_name = tuple[2]
        upstream_total = tuple[3]
        upstream_in = tuple[0]
        upstream_out =tuple[1]


        serial = 1
        index = 0
        upstream = None
        for name in upstream_name:

            upstream_in_util = self.percentage(upstream_in[index], upstream_total[index])
            upstream_out_util = self.percentage(upstream_out[index], upstream_total[index])
            if upstream == None:
                upstream = "<tr class = objects>\n" + "<td>" + str(serial) + "</td>" + "<td>" + str(upstream_name[index]) + "</td>" + "<td>" + str(upstream_total[index]) + "</td>" + "<td>" + str(upstream_in[index]) + "</td>" +"<td>"+ str(upstream_out[index]) + "</td>" + "<td>" + str(upstream_in_util) + "</td>" + "<td>" + str(upstream_out_util) + "</td>\n" +"<td></td>"+ "</tr>\n"
            else:
                upstream = upstream+"<tr class = objects>\n" + "<td>" + str(serial) + "</td>" + "<td>" + str(upstream_name[index]) + "</td>" + "<td>" + str(upstream_total[index]) + "</td>" + "<td>" + str(upstream_in[index]) + "</td>" +"<td>"+ str(upstream_out[index]) + "</td>" + "<td>" + str(upstream_in_util) + "</td>" + "<td>" + str(upstream_out_util) + "</td>\n" +"<td></td>"+ "</tr>\n"
            serial += 1
            index +=1
        return upstream



def total_bw(bw_list):
    total = 0
    for bw in bw_list:
        if bw == None:
            continue
        else:
            total = total+bw
    return total




########################################################################

#enterprise Client

########################################################################
filename ="IC-BW-Report.xlsx"

ent_client_start_row = 11
ent_client = excel_parcer(filename,str(ent_client_start_row),'500')
lsp_start_cell = ent_client.make_list_client()[1]
ent_client_list = ent_client.make_list_client()[0]
ent_client_bw = ent_client.make_list_bw()
ent_total = total_bw (ent_client_bw)
ent_client_remarks = ent_client.make_list_remarks(len(ent_client_list)+10,0)
ent_count=len(ent_client_list)
ent_html = ent_client.make_html(ent_count,ent_client_list,ent_client_bw,ent_client_remarks)

########################################################################

#LSP Client

########################################################################

lsp_client = excel_parcer(filename,lsp_start_cell,'500')
ggc_start_cell = lsp_client.make_list_client()[1]
lsp_client_list = lsp_client.make_list_client()[0]
lsp_client_bw = lsp_client.make_list_bw()
lsp_total = total_bw (lsp_client_bw)
lsp_client_remarks = lsp_client.make_list_remarks(len(lsp_client_list),ent_count+ent_client_start_row) #Enterprise start row number = 11
#print(lsp_client_remarks)
lsp_count = len(lsp_client_list)
lsp_html = lsp_client.make_html(lsp_count,lsp_client_list,lsp_client_bw,lsp_client_remarks)

########################################################################

#GGC Client

########################################################################

ggc_client = excel_parcer(filename,ggc_start_cell,'500')
ggc_client_list = ggc_client.make_list_client()[0]
ggc_client_bw = ggc_client.make_list_bw()
ggc_total = total_bw (ggc_client_bw)
ggc_client_remarks = ggc_client.make_list_remarks(len(ggc_client_list), lsp_count+int(lsp_start_cell))
ggc_count = len(ggc_client_list)
ggc_html = lsp_client.make_html(ggc_count,ggc_client_list,ggc_client_bw,ggc_client_remarks)


########################################################################

#Upstream Bandwidth

########################################################################

upstream = excel_parcer(filename,"5",'6')
upstream_list = upstream.upstream()
up_html = upstream.upstream_html(upstream_list)




########################################################################

#Creating HTML

########################################################################


from datetime import date, timedelta

yesterday = date.today() - timedelta(days=1)
date = yesterday.strftime('%d/%m/%y')

name = ""#input("Please Enter Shift Engineer Name: ")

fname = "table.html"
fhandle = open(fname, "w")
mail_body = """
<html>
    <head>
        <style>
            tr, td, th { border: 1px solid black;
                         text-align: center;
                         padding: 15px;
                          }
            #date {
            width: 10%;
            }


            .table_head {
            background: #a3a375;
            color: #800000;
            font-size: 120%;
            }

            #report_head, #date {
            font-size:150%;
            color: #800000;
            background : #999966
            }
            #date {
            font-size:100%;
            }
            .objects {
            background : #d6d6c2
            }
            .head {
            background-color:#c1c1a4;
            }
            .total {
            font-weight: bold;
            background : #d6d6c2
            }
        </style>
    </head>
    <body>
        <br>
        <p> Dear Sir, <br>
        <br>
        Kindly find the attached mail for InterCloud Bandwidth report:
        </p>
        <table>
            <thead>
                <tr>
                    <th id=date>"""+date+"""</th>
                    <th id=report_head colspan ="7"> InterCloud Bandwidth Report</th>
                </tr>

                <tr>
                    <th colspan="8" class ="Table_head">Upstream</th>
                </tr>

                <tr class=head>
                    <th>Sr. Number</th>
                    <th>Upstream Name</th>
                    <th>Total Capacity(Mbps)</th>
                    <th>Highest Bandwidth IN(Mpbs)</th>
                    <th>Highest Bandwidth Out(Mpbs)</th>
                    <th>Utilization IN (%) </th>
                    <th>Utilization OUT (%) </th>
                    <th>Remarks</th>
                </tr>

                """+up_html+"""

                <tr>
                    <th colspan="8" class ="Table_head">Enterprise Client</th>
                </tr>
                <tr class=head>
                    <th>Sr. Number</th>
                    <th>Name</th>
                    <th>Bandwidth</th>
                    <th colspan='5'>Remarks</th>
                </tr>
            </thead>
            <tbody>""" + ent_html + """
            <tr class = 'total'>
                <td colspan ='2'>Total: </td>
                <td>"""+ str(ent_total) + """</td>
                <td colspan ='5'> </td>
            </tr>
            </tbody>
            <thead>
                <tr>
                    <th colspan="8" class ="Table_head">LSP Client</th>
                </tr>
                <tr class=head>
                    <th>Sr. Number</th>
                    <th>Name</th>
                    <th>Bandwidth</th>
                    <th colspan='5'>Remarks</th>
                </tr>
            </thead>
            <tbody>""" + lsp_html + """
            <tr class = 'total'>
                <td colspan ='2'>Total: </td>
                <td>"""+ str(lsp_total) + """</td>
                <td colspan ='5'> </td>
            </tr>

            </tbody>
            <thead>
                <tr>
                    <th colspan="8" class ="Table_head">GGC Client</th>
                </tr>
                <tr class=head>
                    <th>Sr. Number</th>
                    <th>Name</th>
                    <th>Bandwidth</th>
                    <th colspan='5'>Remarks</th>
                </tr>
            </thead>
            <tbody>""" + ggc_html + """

            <tr class = 'total'>
                <td colspan ='2'>Total: </td>
                <td>"""+ str(ggc_total) + """</td>
                <td colspan ='5'> </td>
            </tr>

            </tbody>
        </table>
        <p>Regards,</p>
        <br>
        <p>""" + name + """</p>
        <p><b>Core Network || AS58923</b></p>
        <p><b>Internet Service Provider (ISP) & IP Telephony Service Provider (IPTSP)</b></p>
        <p>House: Ga-30/G, Pragati Sarani, Shahjadpur, Gulshan-2, Dhaka – 1212, Bangladesh</p>
        <p><b>phone:</b>  +8809638383838; Ext: 8426, 8427</p>
    </body>
</html>"""

#print(mail_body)
fhandle.write(mail_body)
fhandle.close()


########################################################################

#Creating Tunnel and Sending Mail

########################################################################

import email, smtplib, ssl

from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

sender_email = "core@brilliant.com.bd"
receiver_email = ["mosiur.rahman@novocom-bd.com","mosiur.rahman@brilliant.com.bd"]
caption = ["core@brilliant.com.bd"]
password = "ipcore#@!"

message = MIMEMultipart()
message["Subject"] = "multipart test"
message["From"] = sender_email
message["To"] = ','.join(receiver_email)
message["Cc"] = ','.join(caption)


# Turn these into plain/html MIMEText objects
#part1 = MIMEText(text, "plain")
part2 = MIMEText(mail_body, "html")

# Add HTML/plain-text parts to MIMEMultipart message
# The email client will try to render the last part first
#message.attach(part1)
message.attach(part2)


filename = "IC-BW-Report.xlsx"

with open(filename, "rb") as attachment:
    # Add file as application/octet-stream
    # Email client can usually download this automatically as attachment
    part = MIMEBase("application", "octet-stream")
    part.set_payload(attachment.read())

# Encode file in ASCII characters to send by email
encoders.encode_base64(part)

# Add header as key/value pair to attachment part
part.add_header(
    "Content-Disposition",
    f"attachment; filename= {filename}",
)

# Add attachment to message and convert message to string
message.attach(part)
text = message.as_string()


# Create secure connection with server and send email
context = ssl.create_default_context()
with smtplib.SMTP_SSL("mta.brilliant.com.bd", 465, context=context) as server:
    server.login(sender_email, password)
#    server.sendmail(
#        sender_email, receiver_email,text
#    )
